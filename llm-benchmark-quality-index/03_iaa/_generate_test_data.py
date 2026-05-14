"""Generate synthetic evaluator scoring Excel files for development testing.

This script was used during initial development to validate the IAA pipeline
with synthetic data before real evaluator files were available. It remains in
the repository as a development utility:
- Running this script will OVERWRITE the three real evaluator files in
  02_scoring/raw_scores/ with synthetic data. Do not run it accidentally.
- Useful only for: (a) regenerating synthetic data to test changes to
  compute_iaa.py, or (b) demonstrating the pipeline end-to-end without
  exposing real evaluator data.

To restore real evaluator data after running this script, retrieve the
original Excel files from git history.
"""
import shutil
from pathlib import Path
import numpy as np
from openpyxl import load_workbook

np.random.seed(42)

TEMPLATE = Path(__file__).parent.parent / "docs" / "BQAI_scoring_template_ES.xlsx"
OUTPUT_DIR = Path(__file__).parent.parent / "02_scoring" / "raw_scores"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# "Ground truth" scores per benchmark, with noise per evaluator
np.random.seed(42)
n_benchmarks = 30
n_dimensions = 7
TRUE_SCORES = np.clip(np.random.uniform(0.4, 0.95, size=(n_benchmarks, n_dimensions)), 0, 1)

EVALUATORS = ["Evaluator1", "Evaluator2", "Evaluator3"]
NOISE_STD = 0.06  # realistic evaluator noise

# Read template to find which rows have actual benchmark data
wb = load_workbook(TEMPLATE)
ws = wb["Scoring"]
benchmark_rows = []
for row in range(7, 45):
    name = ws.cell(row=row, column=2).value
    if name and isinstance(name, str) and "benchmark" not in name.lower():
        benchmark_rows.append(row)

print(f"Found {len(benchmark_rows)} benchmark rows in template")
assert len(benchmark_rows) == 30, f"Expected 30, got {len(benchmark_rows)}"

for i, name in enumerate(EVALUATORS):
    out_path = OUTPUT_DIR / f"BQAI_scoring_{name}.xlsx"
    shutil.copy(TEMPLATE, out_path)

    wb = load_workbook(out_path)
    ws = wb["Scoring"]
    ws.cell(row=2, column=3, value=name)
    ws.cell(row=3, column=3, value="2026-05-11")

    np.random.seed(42 + i)
    noise = np.random.normal(0, NOISE_STD, size=(n_benchmarks, n_dimensions))
    scores = np.clip(TRUE_SCORES + noise, 0, 1)

    for b_idx, row in enumerate(benchmark_rows):
        for d_idx in range(n_dimensions):
            ws.cell(row=row, column=6 + d_idx, value=round(scores[b_idx, d_idx], 2))

    wb.save(out_path)
    print(f"Generated: {out_path}")

print("\nDone. Run: python 03_iaa/compute_iaa.py")
